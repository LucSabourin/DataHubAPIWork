# Python Standard Library
from io import BytesIO

# OpenPyXL
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl import load_workbook

# Messages
from messages import Messages

# Worksheet
from datatables.xl_worksheet import XLWorksheet


class XLWorkbook:
    """
    Class to scan an Excel Workbook and expose a list of the Worksheet Names
    and a dictionary indexed by those names exposing an openpyxl Worksheet object for each
    worksheet present in the Excel Workbook.
    """

    message_templates = {
        'could_not_load_data_only_into_openpyxl': ('Could not load data from filename {0} as a workbook (xls, xlsx, xlsm).', 99),
        'could_not_load_data_and_formulas_into_openpyxl': ('Could not load data and formulas from filename {0} as a workbook (xls, xlsx, xlsm).', 99),
    }

    def __init__(self, file: BytesIO, filename: str) -> None:
        self.binary = file
        self.data_only = None
        self.with_formulas = None
        self.filename = filename
        self.sheet_names = []
        self.messages = Messages(XLWorkbook.message_templates)
        self.worksheets = {}

    def _open_workbook(self, formula_values: bool) -> tuple:
        """
        Opens the file as an openpyxl Workbook, reporting success or failure, and appending an error message on failure.
        """
        succeeded = True
        workbook = None
        try:
            workbook = load_workbook(self.binary, data_only=formula_values)
        except:
            # We are intentionally absorbing the exception, so that we can report the message and
            # keep executing.
            succeeded = False
            if formula_values:
                self.messages.add_message('could_not_load_data_only_into_openpyxl', self.filename)
            else:
                self.messages.add_message('could_not_load_data_and_formulas_into_openpyxl', self.filename)
        return succeeded, workbook

    def _has_place_holder_header_on_a_worksheet(self) -> bool:
        """
        Scans the worksheets to see if any of them had a place-holder row inserted
        as a header and returns True at least one of the worksheets has a place-holder inserted,
        otherwise returns False.
        """
        for sheet in self.worksheets.values():
            if sheet.metadata['has_place_holder_header']:
                return True
        return False

    def _analyze_worksheet(self, sheet_name: str, workbook_data_only: Workbook, workbook_with_formulas: Workbook):
        """
        Builds worksheet metadata from OpenPyXL worksheet objects.
        """
        data_only = workbook_data_only[sheet_name]
        with_formulas = workbook_with_formulas[sheet_name]
        worksheet = XLWorksheet(sheet_name, data_only, with_formulas)
        worksheet.build_metadata()
        self.worksheets[sheet_name] = worksheet

    def analyze_workbook(self) -> bool:
        """
        Loads the Excel workbook in both data only and formulas mode, and then builds the metadata.

        If while building the metadata, place-holders were inserted into the raw data, updates our
        binary image to one containing the place-holders.

        Returns False if an error occurred loading the data only or formulas worksheets, and True
        if it successfully loads ands creates metadata for the Excel workbook.
        """
        succeeded, data_only = self._open_workbook(True)
        if not succeeded:
            return False
        
        self.data_only = data_only
        succeeded, with_formulas = self._open_workbook(False)
        if not succeeded:
            return False
        
        self.with_formulas = with_formulas
        self.sheet_names = list(data_only.sheetnames)
        for sheet_name in self.sheet_names:
            self._analyze_worksheet(sheet_name, data_only, with_formulas)

        # If we added any place-holder headers on a worksheet,
        # we have to update the binary we are going to pass to pandas
        if self._has_place_holder_header_on_a_worksheet:
            self.binary = BytesIO(save_virtual_workbook(data_only))
        return True

