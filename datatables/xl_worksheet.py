# OpenPyXL dependencies
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# Json
import json

# Models dependencies
import datatables.scan.scan_excel as scan
from datatables.scan.scan_generic import duplicate_headers
from messages import Messages


class XLWorksheet:
    """
    Class that acts as a container for an Excel Worksheet.

    Contains openpyxl worksheet objects and metadata about the loaded worksheet.

    Metadata Keys
    -------------
    The metadata built to describe the worksheet contains the following keys:

    Keys that are always present:
    name - str - the worksheet name
    source_type - str - describes the type of raw data source
    has_place_holder_header - bool - indicates if the openpyxl worksheet was modified by an inserting empty first
        row where the header would normally be.
    first_row - integer - the first row of the worksheet that contains data
    last_row - integer - the last row of the worksheet that contains data
    first_column - integer - the first column of the worksheet that contains data
    last_column - integer - the last column of the worksheet that contains data
    headers - list - the list of column names
    field_formats - dict - description of the field (column) data types
    field_header_row - integer - the row that contains the header
    """

    message_templates = {
        'top_rows_empty': ('The top {1} rows of worksheet {0} are empty and so have been omitted.', 0),
        'left_columns_empty': ('The leftmost {1} columns of worksheet {0} are empty and so have been omitted.', 0),
        'merged_cells': ('Cells in range {1} found in worksheet {0} have been unmerged, so contents have been stored in cell {2} and all other cells in the range are empty.',0),
        'missing_headers': ('There are missing headers in columns {1} of worksheet {0}.', 1),
        'duplicate_headers': ('The header {1} is duplicated {2} times in columns {3} in worksheet {0}.', 1),
        'hierarchical_headers': ('The hierarchical headers {1} in worksheet {0} have been combined into a single header, with <.> separating each layer (top layer on the left and bottom layer on the right).',1),
        'hidden_rows': ('Hidden rows {1} in worksheet {0} have been unhidden, so their contents have been included.', 2),
        'formula_error': ('Formula {1} in cell {3} of worksheet {0} resulted in error {2}.', 3),
    }

    def __init__(self, name: str, data_only: Worksheet, with_formulas: Worksheet, ):
        self.metadata = {
            'has_place_holder_header': False,
            'source_type': 'Excel Worksheet',
            'name': name,
        }
        self.sheet = data_only
        self.formula_sheet = with_formulas
        self.messages = Messages(XLWorksheet.message_templates)

    def build_metadata(self):
        """
        Scans the openpyxl worksheet and builds the metadata dictionary
        """
        self._build_size_metadata()
        self._emit_empty_rows_cols_messages()
        self._emit_merged_cells_messages()
        self._build_header_metadata()
        self._emit_hidden_rows_messages()
        self._emit_formulas_error_messages()
        # Jsonifies metadata.
        self.metadata = json.loads(json.dumps(self.metadata))


    def _emit_formulas_error_messages(self):
        """
        Adds error messages to the messages object if there are any formula errors.
        """
        # Formulas
        formulas = scan.sheet_formulas(
            sheet=self.formula_sheet,
            min_row_data=self.metadata['first_row'],
            max_row=self.metadata['last_row'],
            min_col=self.metadata['first_column'],
            max_col=self.metadata['last_column']
        )
        if formulas is not None:
            # Formula Errors
            for cell_ref, formula in formulas.items():
                if scan.sheet_check_formula_error(
                    sheet=self.sheet,
                    row_index=formula[0],
                    col_index=formula[1]
                ) is True:
                    self.messages.add_message(
                        'formula_error',
                        self.metadata['name'],
                        formula[2],
                        self.sheet.cell(formula[0],
                        formula[1]).value,
                        cell_ref
                    )

    def _emit_hidden_rows_messages(self):
        """Adds error messages to the messages object if there are any formula errors."""
        # Hidden Rows
        hidden_rows = scan.sheet_hidden_rows(
            sheet=self.sheet,
            min_row=self.metadata['first_row'],
            max_row=self.metadata['last_row']
        )
        if len(hidden_rows) > 0:
            self.messages.add_message('hidden_rows', self.metadata['name'], hidden_rows)

    def _build_header_metadata(self):
        """
        Builds metadata related to the worksheet header
        """
        # Header info
        field_header_row = scan.sheet_field_headers_row(
            sheet=self.sheet,
            min_row=self.metadata['first_row'],
            max_row=self.metadata['last_row'],
            min_col=self.metadata['first_column'],
            max_col=self.metadata['last_column']
        )
        if field_header_row != -1:
            # Has Headers in file
            self._build_header_metadata_from_row(field_header_row)
        else:
            # Does Not Have Headers in file
            self._build_header_metadata_no_header()

    def _build_header_from_row(self, field_header_row):
        self.metadata['field_header_row'] = field_header_row
        headers = scan.sheet_field_headers(
            sheet=self.sheet,
            header_row=field_header_row,
            min_col=self.metadata['first_column'],
            max_col=self.metadata['last_column']
        )
        return headers

    def _build_header_metadata_no_header(self):
        """
        Constructs a generic header metadata and inserts an empty place-holder header row in the worksheet.
        """
        # Add empty row above data, so it does not get cut off
        if self.metadata['first_row'] == 1:
            self.sheet.insert_rows(1)
            self.metadata['has_place_holder_header'] = True
            self.metadata['field_header_row'] = 1

        headers = self._build_header_metadata_create_placeholders()
        self._emit_missing_headers_messages(headers)

        # Field formats
        self._build_field_formats_metadata(self.metadata['first_row'])
        return headers

    def _emit_missing_headers_messages(self, headers: list):
        """
        Adds error messages to the messages object if there are missing headers.
        """
        # Missing Field Headers
        missing_headers = [num + self.metadata['first_column'] for num in range(0, len(headers))]
        self.messages.add_message(
            'missing_headers',
            self.metadata['name'],
            [get_column_letter(pos) for pos in missing_headers]
        )

    def _build_header_metadata_create_placeholders(self):
        """
        Constructs generic header metadata.
        """
        # Base Header Info
        headers = []
        for num in range(self.metadata['first_column'], self.metadata['last_column'] + 1):
            headers.append('Unnamed: ' + str(num))
        self.metadata['headers'] = headers
        return headers

    def _build_field_formats_metadata(self, field_header_row):
        """
        Constructs field formats metadata.
        """
        # Field Formats
        self.metadata['field_formats'] = scan.sheet_field_formats(
            sheet=self.sheet,
            min_row_data=field_header_row,
            min_col=self.metadata['first_column'],
            max_col=self.metadata['last_column']
        )

    def _build_header_metadata_from_row(self, field_header_row):
        """
        Constructs header metadata from the header row in the worksheet.
        """
        # Base Header Info.
        self.metadata['field_header_row'] = field_header_row
        headers = scan.sheet_field_headers(
            sheet=self.sheet,
            header_row=field_header_row,
            min_col=self.metadata['first_column'],
            max_col=self.metadata['last_column']
        )
        # Field Formats
        self._build_field_formats_metadata(field_header_row + 1)

        self._emit_duplicate_headers_messages(headers)

        # Hierarchical Field Headers
        hierarchical_headers = scan.sheet_hierarchical_fields(
            sheet=self.sheet,
            min_row=self.metadata['first_row'],
            header_row=field_header_row,
            min_col=self.metadata['first_column'],
            max_col=self.metadata['last_column']
        )
        if len([header for header in hierarchical_headers if '.' in str(header or '')]) > 0:
            headers = hierarchical_headers
            self.messages.add_message(
                'hierarchical_headers',
                self.metadata['name'],
                hierarchical_headers
            )

        self._emit_missing_field_headers_messages(headers)
        self.metadata['headers'] = headers

    def _emit_duplicate_headers_messages(self, headers):
        """
        Adds error messages to the messages object if there are duplicate headers.
        """
        # Duplicate Field Headers
        dupe_headers = duplicate_headers(
            headers=headers,
            min_col=self.metadata['first_column']
        )
        if dupe_headers is not None:
            for header, positions in dupe_headers:
                self.messages.add_message(
                    'duplicate_headers',
                    self.metadata['name'],
                    header,
                    len(positions),
                    positions
                )

    def _emit_missing_field_headers_messages(self, headers):
        """
        Adds error messages to the messages object if there are headers are missing fields.
        """
        # Missing Field Headers
        if None in headers:
            missing_headers = [num + self.metadata['first_column'] for num, header in enumerate(headers) \
                                if header is None]
            self.messages.add_message(
                'missing_headers',
                self.metadata['name'],
                [get_column_letter(pos) for pos in missing_headers]
            )

    def _emit_merged_cells_messages(self):
        """
        Adds error messages to the messages object if the worksheet has merged cells.
        """
        # Merged Cells
        merged_cells = scan.sheet_merged_cells(self.sheet)
        if merged_cells is not None:
            for cell_range, info in merged_cells.items():
                self.messages.add_message(
                    'merged_cells',
                    self.metadata['name'],
                    cell_range,
                    info[0]
                )

    def _emit_empty_rows_cols_messages(self):
        """
        Adds error messages to the messages object if the worksheet has one or more initial empty rows.
        """
        # Number of empty rows to the top of data
        if self.metadata['first_row'] > 1:
            self.messages.add_message(
                'top_rows_empty',
                self.metadata['name'],
                self.metadata['first_row'] - 1
            )

        # Number of empty columns to the left of data
        if self.metadata['first_column'] > 1:
            self.messages.add_message(
                'left_columns_empty',
                self.metadata['name'],
                self.metadata['first_column'] - 1
            )

    def _build_size_metadata(self):
        """
        Constructs the metadate related to the row and column sizes of the worksheet.
        """
        # Min/Max Row and Column Indices
        self.metadata['first_row'] = scan.sheet_min_row(self.sheet)
        self.metadata['last_row'] = scan.sheet_max_row(self.sheet)
        self.metadata['first_column'] = scan.sheet_min_col(self.sheet)
        self.metadata['last_column'] = scan.sheet_max_col(self.sheet)
